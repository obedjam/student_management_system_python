import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import font as tkfont
from tkinter.filedialog import asksaveasfile
import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import date
import xlrd
from functools import partial 
from operator import ne
import shutil
import os
from PIL import Image, ImageDraw, ImageFont
from firebase_admin import credentials, initialize_app, storage
import urllib.request

def start():
    window.destroy()
    window3 = Tk()
    window3.title("Student Management System")
    window3.iconbitmap('data/icon.ico')
    width, height = window3.winfo_screenwidth(), window3.winfo_screenheight()
    window3.geometry('%dx%d+0+0' % (width,height))
    window3.state("zoomed")
    window3.configure(bg="white")
    bg = PhotoImage(file="data/bg.png")
    rp = PhotoImage(file="data/tmp_reciept.png")
    my_canvas = Canvas(window3, width=800, height=500)
    my_canvas.pack(fill="both", expand=True)
    my_canvas.create_image(0,0, image=bg, anchor="nw")
    txtfont=tkfont.Font(size=10)
    txtfont1=tkfont.Font(size=20,family="Times New Roman")
    txtfont2=tkfont.Font(size=14,family="Times New Roman")
    txtfont3=tkfont.Font(size=12)
    txtfont4=tkfont.Font(size=14)
    txtfont5=tkfont.Font(size=6)
    def connect():
        try:
            urllib.request.urlopen('http://google.com') #Python 3.x
            return True
        except:
            return False
    def update_data():
        global app_init
        if connect():
            # Init firebase with your credentials
            cred = credentials.Certificate("firebase.json")
            if not app_init:
                initialize_app(cred, {'storageBucket': 'firebase.com'})
                app_init = True
            # Put your local file path 
            fileName = "./data/studentlist.xlsx"
            bucket = storage.bucket()
            blob = bucket.blob(fileName)
            blob.upload_from_filename(fileName)

            fileName = "./data/feestatus.xlsx"
            bucket = storage.bucket()
            blob = bucket.blob(fileName)
            blob.upload_from_filename(fileName)

            fileName = "./data/fee.xlsx"
            bucket = storage.bucket()
            blob = bucket.blob(fileName)
            blob.upload_from_filename(fileName)

            fileName = "./data/admissions.xlsx"
            bucket = storage.bucket()
            blob = bucket.blob(fileName)
            blob.upload_from_filename(fileName)
            messagebox.showinfo("Update successful","Data was updated successfully.")
        else:
            messagebox.showwarning("Connection Error","Please check your internet connection.")
        
    def getLatestSheet():
        wb = load_workbook('./data/studentlist.xlsx')
        sheet = wb.sheetnames[len(wb.sheetnames)-1]
        wb.close()
        return sheet
        
    def resize():
        basewidth = 350
        baseheight = 380
        img = Image.open('data/tmp_reciept.png')
        wpercent = (basewidth / float(img.size[0]))
        img = img.resize((basewidth, img.size[1]), Image.LANCZOS)
        hpercent = (baseheight / float(img.size[1]))
        img = img.resize((img.size[0], baseheight), Image.LANCZOS)
        img.save('data/sreciept.png')
    def getdata():
        global batch
        df=pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False, sheet_name=str(batch))
        df_rows = df.to_numpy().tolist()
        return df_rows
            

    def adid():
        df_rows = []
        df = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=None,header=None,index_col=False)
        for frames in df:
            df_rows = df_rows+df[frames].to_numpy().tolist()
        val1_prev=0
        val1=0
        for value in df_rows:
            val=value[26]
            if val != "AdmissionID":
                print(val[4:])
                if int(val[4:])>=val1_prev:
                    val1=int(val[4:])
                val1_prev=val1
            
        newid="ID"+str(val1+1).zfill(6)
        return newid
    def save_file(img):
        file=asksaveasfile(mode='w', initialfile="reciept.png", defaultextension=".png",filetypes=[("All Files","*,*")])
        path=os.path.abspath(file.name)
        img.save(path)

    def assign_roll():
        cls_list=[]
        sheet = getLatestSheet()
        df=pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False, sheet_name=str(sheet))
        df_rows = df.to_numpy().tolist()
        for data in df_rows:
            if data[18] not in cls_list:
                cls_list.append(data[18])
        for index in cls_list:
            cls_num=[]
            for data in df_rows:
                if index == data[18]:
                    cls_num.append(str(data[0])+" "+str(data[1]))
            cls_num.sort()
            roll=1
            for item in cls_num:
                i=1
                wb = load_workbook('./data/studentlist.xlsx')
                ws=wb[sheet]
                for data in df_rows:
                    if item == str(data[0])+" "+str(data[1]) and index==data[18]:
                        ws['T'+str(i+1)]=roll
                        wb.save('./data/studentlist.xlsx')
                        roll=roll+1
                    i=i+1
                wb.close()
            
            
        
    def state1():
        global state
        global batch
        global df_rows
        batch=getLatestSheet()
        df_rows=getdata()
        state=1
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state2():
        global state
        state=2
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state3():
        global state
        state=3
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state4():
        global state
        global batch
        global df_rows
        batch=getLatestSheet()
        df_rows=getdata()
        state=4
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state5():
        global state
        state=5
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def form():
        global months
        global df_rows
        global batch
        months=[]
        v = StringVar()
        v.set("Tags: Students")
        u = StringVar()
        u.set("For the month of")
        v1 = StringVar()
        v1.set("Count: ")
        v2 = StringVar()
        v2.set("Tags: Entries")
        var1 = tk.IntVar()
        def check_fee(adid):
            feefile = pd.read_excel('./data/feestatus.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
            feedata = feefile.to_numpy().tolist()
            date=today.strftime("%d/%m/%Y")
            subdate=date.split("/")
            if(batch != subdate[2]):
                subdate[1]=11
            for value in feedata:
                if value[0]==adid:
                    for i in range(1,int(subdate[1])+1):
                        if value[i]=="not paid":
                            return 0
            return 1

        def addmonth():
            global months
            if variable.get()!="Month":
                months.append(variable.get())
                u.set(u.get()+", "+variable.get())
            return
        def to_excel(admission,uniform,other,books,des):
            sheet=today.strftime("%Y")
            wb = load_workbook('./data/studentlist.xlsx')
            if sheet not in wb.sheetnames:
                ws=wb.create_sheet(sheet)
                ws.append(['First Name','Last name','Gender','Age','Father\'s Name','Mother\'s Name','Gaurdian\'s Name','DOB','Religion','Address','Father\'s Contact No.','Mother\'s Contact No.','Identity Mark','Blood Group','Community','Tribe/Caste','Father\'s Occupation','Mother\'s Occupation','Class','Roll No.','Previous School','Result','Percentage','Prev. Attendance','Admission Date','Last Fee Payment Date','AdmissionID'
    ])
            wb.save('./data/studentlist.xlsx')
            wb.close()
                
            wb = load_workbook('./data/admissions.xlsx')
            if sheet not in wb.sheetnames:
                ws=wb.create_sheet(sheet)
                ws.append(['AdmissionID','Admission fee','Admission Date','Books and Statonary fee','Uniform fee','Other','Description'])
            wb.save('./data/admissions.xlsx')
            wb.close()

            wb = load_workbook('./data/feestatus.xlsx')
            if sheet not in wb.sheetnames:
                ws=wb.create_sheet(sheet)
                ws.append(['AdmissionID','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Admission Date'])
            wb.save('./data/feestatus.xlsx')
            wb.close()

            wb = load_workbook('./data/fee.xlsx')
            if sheet not in wb.sheetnames:
                ws=wb.create_sheet(sheet)
                ws.append(['AdmissionID','Fee Amount','Books and Statonary fee','Uniform fee','Other','Description','Payment Date'])
            wb.save('./data/fee.xlsx')
            wb.close()
            
            newid=adid()
            entry=[entry1.get(),entry2.get(),variable2.get(),entry4.get(),entry5.get(),entry6.get(),entry7.get(),entry8.get(),entry9.get(),entry10.get(),entry11.get(),entry12.get(),entry18.get(),entry13.get(),entry25.get(),variable.get(),entry15.get(),entry16.get(),(entry17.get()).upper(),"0",entry19.get(),variable1.get(),entry22.get(),entry23.get(),today.strftime("%d/%m/%Y")," ",newid]
            entryad=[newid,admission,today.strftime("%d/%m/%Y"),books,uniform,other,des]
            for value in entryad:
                if value=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return False
            wb = load_workbook('./data/studentlist.xlsx')
            ws = wb[sheet]
            ws.append(entry)
            wb.save('./data/studentlist.xlsx')
            wb.close()
            wb = load_workbook('./data/admissions.xlsx')
            ws = wb[sheet]
            ws.append(entryad)
            wb.save('./data/admissions.xlsx')
            wb.close()
            wb = load_workbook('./data/feestatus.xlsx')
            ws = wb[sheet]
            entryfee=[newid,"not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid",today.strftime("%d/%m/%Y")]
            ws.append(entryfee)
            wb.save('./data/feestatus.xlsx')
            wb.close()
            messagebox.showinfo("Entry successful","Student data has been saved successfully.\nAdmission ID: "+newid)
            return True
        
        def to_excel1():
            global months
            sheet=getLatestSheet()
            batch=sheet
            df_rows=getdata()
                
            wb = load_workbook('./data/studentlist.xlsx')
            ws = wb[sheet]
            i=1
            for data in df_rows:
                if str(data[26]).lower()==str(entry1.get()).lower():
                    ws['Z'+str(i+1)]=today.strftime("%d/%m/%Y")
                    wb.save('./data/studentlist.xlsx')
                    wb.close()
                    break
                i=i+1
                
                    
             
            entry=[(entry1.get()).upper(),entry4.get(),entry5.get(),entry8.get(),entry6.get(),entry7.get(1.0, "end-1c"),today.strftime("%d/%m/%Y")]
            wb = load_workbook('./data/fee.xlsx')
            ws = wb[sheet]
            ws.append(entry)
            wb.save('./data/fee.xlsx')
            wb.close()
            i=1
            for data in df_rows:
                if str(data[26]).lower()==str(entry1.get()).lower():
                    for item in months:
                            wb = load_workbook('./data/feestatus.xlsx')
                            ws = wb[sheet]
                            ws[OptionList[item]+str(i+1)]=today.strftime("%d/%m/%Y")
                            wb.save('./data/feestatus.xlsx')
                            wb.close()
                    break
                i=i+1
            messagebox.showinfo("Entry successful","Fee Entry has been saved successfully.")
            
            return True

            
        def to_excel2(admission,uniform,other,books,des):
            sheet=today.strftime("%Y")      
            wb = load_workbook('./data/admissions.xlsx')
            if sheet not in wb.sheetnames:
                ws=wb.create_sheet(sheet)
                ws.append(['AdmissionID','Admission fee','Admission Date','Books and Statonary fee','Uniform fee','Other','Description'])
            wb.save('./data/admissions.xlsx')
            wb.close()

            wb = load_workbook('./data/feestatus.xlsx')
            if sheet not in wb.sheetnames:
                ws=wb.create_sheet(sheet)
                ws.append(['AdmissionID','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Admission Date'])
            wb.save('./data/feestatus.xlsx')
            wb.close()

            wb = load_workbook('./data/fee.xlsx')
            if sheet not in wb.sheetnames:
                ws=wb.create_sheet(sheet)
                ws.append(['AdmissionID','Fee Amount','Books and Statonary fee','Uniform fee','Other','Description','Payment Date'])
            wb.save('./data/fee.xlsx')
            wb.close()
            
            tmp = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False)
            entry=[(entry4.get()).upper(),admission,today.strftime("%d/%m/%Y"),books,uniform,other,des]
            for value in entry:
                if value=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return False

            wb = xlrd.open_workbook('data/studentlist.xlsx')
            oldsheet = wb.sheet_by_name(str(int(sheet)-1))
            oldsheet.cell_value(0, 0)
            for i in range(1,oldsheet.nrows):
                if str(oldsheet.cell_value(i, 26)).lower()==str(entry4.get()).lower():
                    wb = load_workbook('./data/studentlist.xlsx')
                    ws = wb[sheet]
                    tmp_row = oldsheet.row_values(i)
                    tmp_row[24]=today.strftime("%d/%m/%Y")
                    tmp_row[20]="My School"
                    tmp_row[21]=variable.get()
                    tmp_row[22]=entry22.get()
                    tmp_row[23]=entry23.get()
                    tmp_row[18]=entry21.get()
                    ws.append(tmp_row)
                    wb.save('./data/studentlist.xlsx')
                    wb.close()
                    wb = load_workbook('./data/admissions.xlsx')
                    ws = wb[sheet]
                    ws.append(entry)
                    wb.save('./data/admissions.xlsx')
                    wb.close()
                    wb = load_workbook('./data/feestatus.xlsx')
                    ws = wb[sheet]
                    entryfee=[(entry4.get()).upper(),"not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid",today.strftime("%d/%m/%Y")]
                    ws.append(entryfee)
                    wb.save('./data/feestatus.xlsx')
                    wb.close()
                    break
                else:
                    if i==oldsheet.nrows:
                        messagebox.showwarning("Oops!","Student data not found..")
                        return False
                    else:
                        continue   
            messagebox.showinfo("Entry successful","Admission has been renewed successfully.")
            return True
        def rd_excel():
            global df_rows
            global count
            global batch
            global tags
            count=0
            currentdate=date.today()
            if variable.get()=="Fee Status":
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    for value in df_rows:
                        if entry1.get().lower()=="paid":
                            if check_fee(value[26])== 1:
                                if var1.get()== 1:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                                else:
                                    continue
                            else:
                                if var1.get()== 1:
                                    continue 
                                else:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            messagebox.showwarning("Oops!","Invalid Entry.")
                            return
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    for row in df_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable.get()+"="+entry1.get() not in tags:
                         tags.append("Fee Status="+entry1.get())
                    set_tags(var1.get())
                    entry1.delete(0,'end')
                    variable.set("Select Filter")
                    return

            elif variable.get()=="Batch":
                tags=[]
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    batch=entry1.get()
                    df_rows=getdata()
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    for row in df_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable.get()+"="+entry1.get() not in tags:
                        tags.append("Batch="+entry1.get())
                    set_tags(var1.get())
                    entry1.delete(0,'end')
                    variable.set("Select Filter")
                    return
                    
            elif variable.get()!="Select Filter":
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    txt=entry1.get().lower().split(",")
                    tmp_rows=[]
                    for item in txt:
                        if var1.get()==1:
                            df_rows=df_rows
                        elif var1.get()==0 and ("," in entry1.get()):
                            df_rows = getdata()
                        for value in df_rows:
                            if str(value[OptionList[variable.get()]]).lower()!=item:
                                if var1.get()== 1:
                                    continue
                                else:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                            else:
                                if var1.get()== 1:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                                else:
                                    continue
                        if var1.get()==1:
                            tmp_rows=df_rows
                        elif var1.get()==0:
                            tmp_rows=tmp_rows+df_rows 
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    for row in tmp_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable.get()+"="+entry1.get() not in tags:
                        tags.append(variable.get()+"="+entry1.get())
                    set_tags(var1.get())
                    entry1.delete(0,'end')
                    variable.set("Select Filter")
                    df_rows=tmp_rows
                    return

            else:
                messagebox.showwarning("Oops!","Please select a filter")
                entry1.delete(0,'end')
                return
            return
        def refresh_excel():
            global df_rows
            global count
            global batch
            count=0
            batch=getLatestSheet()
            df_rows = getdata()
            treeview.delete(*treeview.get_children())
            for column in treeview["column"]:
                treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                 treeview_sort_column(treeview, _column, False))
            for row in df_rows:
                if count % 2 == 0:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                else:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                count=count+1
            v.set("Tags: Students")
            v1.set("Count: "+str(count))
            entry1.delete(0,'end')
            return
        def remove_filter():
            global df_rows
            global batch
            global tags
            entry1.delete(0,'end')
            batch=getLatestSheet()
            df_rows = getdata()
            v.set("Tags: Students")
            v1.set("Count: "+str(count))
            if tags:
                tags.pop()
            if not tags:
                refresh_excel()
            else:
                for items in tags:
                    data=items.split("=")
                    variable.set(data[0])
                    entry1.insert(0,data[1])
                    rd_excel() 
                    tags.pop()       
            return
        def treeview_sort_column(tv, col, reverse):
            l = [(tv.set(k, col), k) for k in tv.get_children('')]
            l.sort(reverse=reverse)

        # rearrange items in sorted positions
            for index, (val, k) in enumerate(l):
                tv.move(k, '', index)

        # reverse sort next time
            tv.heading(col, text=col, command=lambda _col=col: \
                         treeview_sort_column(tv, _col, not reverse))

        def set_tags(check):
            global tags
            global state
            if state==1:
                v.set("Tags: Students")
                for items in tags:
                    tag1=items.split("=")
                    if check==1:
                        v.set(v.get()+", "+tag1[0]+" not "+tag1[1])
                    else:
                        v.set(v.get()+", "+tag1[0]+" "+tag1[1])
            if state==4:
                v2.set("Tags: Entries")
                for items in tags:
                    tag1=items.split("=")
                    if check==1:
                        v2.set(v2.get()+", "+tag1[0]+" not "+tag1[1])
                    else:
                        v2.set(v2.get()+", "+tag1[0]+" "+tag1[1])
            return                    

        def log(check):
            def refresh():
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                    menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                return
            global index
            global df_rows
            global cls
            global count
            global tags
            global batch
            i=0
            feeamnt=0
            addamnt=0
            basamnt=0
            uniamnt=0
            othamnt=0
            v1.set("Count: ")
            col=["","Count","Amount","","","",""]
            treeview.delete(*treeview.get_children())
            if variable.get()=="Total":
                if check==1:
                    tags=[]
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                    if i<2:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                    i=i+1
                i=0
                index=0
                treeview["column"] = col
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df_rows=[]
                    df = pd.read_excel('./data/admissions.xlsx', dtype=str, keep_default_na=False,sheet_name=None)
                    for frames in df:
                        df_rows = df_rows+df[frames].to_numpy().tolist()
                        
                    sdf_rows=[]
                    sdf = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=None)
                    for sframes in sdf:
                        sdf_rows = sdf_rows+sdf[sframes].to_numpy().tolist()
                    
                    for value in df_rows:
                        date1=value[2]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        
                        for data in sdf_rows:
                            if data[26]==value[0]:
                                cls_val=data[18]
                        
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                if var1.get()==1:
                                   if (cls_val not in cls or (not cls)):
                                        i=i+1
                                        addamnt=addamnt+int(value[1])
                                        basamnt=basamnt+int(value[3])
                                        uniamnt=uniamnt+int(value[4])
                                        othamnt=othamnt+int(value[5]) 
                                else:
                                    if (cls_val in cls or (not cls)):
                                        i=i+1
                                        addamnt=addamnt+int(value[1])
                                        basamnt=basamnt+int(value[3])
                                        uniamnt=uniamnt+int(value[4])
                                        othamnt=othamnt+int(value[5]) 
                        
                                
                    addcount=i
                    i=0;
                    df_rows=[]
                    df = pd.read_excel('./data/fee.xlsx', dtype=str, keep_default_na=False,sheet_name=None)
                    for frames in df:
                        df_rows = df_rows+df[frames].to_numpy().tolist()

                    sdf_rows=[]
                    sdf = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=None)
                    for sframes in sdf:
                        sdf_rows = sdf_rows+sdf[sframes].to_numpy().tolist()
                        
                    for value in df_rows:
                        date1=value[6]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        
                        for data in sdf_rows:
                            if data[26]==value[0]:
                                cls_val=data[18]
                                
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                if var1.get()==1:
                                    if (cls_val not in cls or (not cls)):
                                        i=i+1
                                        feeamnt=feeamnt+int(value[1])
                                        basamnt=basamnt+int(value[2])
                                        uniamnt=uniamnt+int(value[3])
                                        othamnt=othamnt+int(value[4])   
                                else:
                                    if (cls_val in cls or (not cls)):
                                        i=i+1
                                        feeamnt=feeamnt+int(value[1])
                                        basamnt=basamnt+int(value[2])
                                        uniamnt=uniamnt+int(value[3])
                                        othamnt=othamnt+int(value[4])
                                 
                            
                    feecount=i
                else:
                    df = pd.read_excel('./data/admissions.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    df_rows = df.to_numpy().tolist()
                    
                    sdf = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    sdf_rows = sdf.to_numpy().tolist()
                            
                    for value in df_rows:
                        for data in sdf_rows:
                            if data[26]==value[0]:
                                cls_val=data[18]
                                
                        if var1.get()==1:
                            if (cls_val not in cls or (not cls)):
                                i=i+1
                                addamnt=addamnt+int(value[1])
                                basamnt=basamnt+int(value[3])
                                uniamnt=uniamnt+int(value[4])
                                othamnt=othamnt+int(value[5]) 
                        else:
                            if (cls_val in cls or (not cls)):
                                i=i+1
                                addamnt=addamnt+int(value[1])
                                basamnt=basamnt+int(value[3])
                                uniamnt=uniamnt+int(value[4])
                                othamnt=othamnt+int(value[5]) 
                    addcount=i
                    i=0
                    df = pd.read_excel('./data/fee.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    df_rows = df.to_numpy().tolist()
                    
                    sdf = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    sdf_rows = sdf.to_numpy().tolist()
                            
                    for value in df_rows:
                        for data in sdf_rows:
                            if data[26]==value[0]:
                                cls_val=data[18]
                                
                        if var1.get()==1:
                            if (cls_val not in cls or (not cls)):
                                i=i+1
                                feeamnt=feeamnt+int(value[1])
                                basamnt=basamnt+int(value[2])
                                uniamnt=uniamnt+int(value[3])
                                othamnt=othamnt+int(value[4])
                        else:
                            if (cls_val in cls or (not cls)):
                                i=i+1
                                feeamnt=feeamnt+int(value[1])
                                basamnt=basamnt+int(value[2])
                                uniamnt=uniamnt+int(value[3])
                                othamnt=othamnt+int(value[4])

                    feecount=i
                count=0
                data=[["Admissions",addcount,addamnt],["Fee Entries",feecount,feeamnt],["Books And Stationaries","NA",basamnt],["Uniform","NA",uniamnt],["Other","NA",othamnt]]
                for row in data:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                cls=[]
                return

                    
            elif variable.get()=="Admissions":
                if check==1:
                    tags=[]
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                df = pd.read_excel('./data/admissions.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                
                treeview["column"] = list(df.columns)
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df_rows=[]
                    df = pd.read_excel('./data/admissions.xlsx', dtype=str, keep_default_na=False,sheet_name=None)
                    for frames in df:
                        df_rows = df_rows+df[frames].to_numpy().tolist()
                        
                    for value in df_rows:
                        date1=value[2]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                continue
                            else:
                                df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            df_rows =list(filter(partial(ne, value), df_rows))
                else:
                    df = pd.read_excel('./data/admissions.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    df_rows = df.to_numpy().tolist()
                count=0
                for row in df_rows:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                v1.set("Count: "+str(count))
                return
            elif variable.get()=="Fee Entries":
                if check==1:
                    tags=[]
                index=8
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                
                df = pd.read_excel('./data/fee.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                treeview["column"] = list(df.columns)
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df_rows=[]
                    df = pd.read_excel('./data/fee.xlsx', dtype=str, keep_default_na=False,sheet_name=None)
                    for frames in df:
                        df_rows = df_rows+df[frames].to_numpy().tolist()
                    for value in df_rows:
                        date1=value[6]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                continue
                            else:
                                df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            df_rows =list(filter(partial(ne, value), df_rows))
                else:
                    df = pd.read_excel('./data/fee.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    df_rows = df.to_numpy().tolist()
                for row in df_rows:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                count=0
                v1.set("Count: "+str(count))
                return
            elif variable.get()=="Fee Status":
                if check==1:
                    tags=[]
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                        
                df = pd.read_excel('./data/feestatus.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                treeview["column"] = list(df.columns)
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df_rows=[]
                    df = pd.read_excel('./data/feestatus.xlsx', dtype=str, keep_default_na=False,sheet_name=None)
                    for frames in df:
                        df_rows = df_rows+df[frames].to_numpy().tolist()
                    for value in df_rows:
                        date1=value[13]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                continue
                            else:
                                df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            df_rows =list(filter(partial(ne, value), df_rows))
                else:
                    df = pd.read_excel('./data/feestatus.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    df_rows = df.to_numpy().tolist()
                count=0
                for row in df_rows:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                v1.set("Count: "+str(count))
                return

                    
        def excel_read():
            global df_rows
            global count
            global index
            global batch
            global cls
            global tags
            count=0
            if entry3.get()=="":
                messagebox.showwarning("Oops!","Some fields are still empty.")
                return
            elif variable1.get()=="Select Filter":
                messagebox.showwarning("Oops!","Please Select a filter.")
                return
            if variable1.get()=="Batch":
                wb = load_workbook('./data/studentlist.xlsx')
                if entry3.get()in wb.sheetnames:
                    wb.close()
                    entry1.delete(0,'end')
                    entry2.delete(0,'end')
                    entry1.insert(0,"1/1/"+entry3.get())
                    entry2.insert(0,"31/12/"+entry3.get())
                    batch=entry3.get()
                    if variable1.get()+"="+entry3.get() not in tags:
                        tags.append(variable1.get()+"="+entry3.get())
                    
                    if variable1.get()!="Select Filter" and entry3.get()!="":
                        set_tags(var1.get())
                    entry3.delete(0,'end')
                    variable1.set("Select Filter")
                    log(0)
                wb.close()
                return


            elif variable.get()=="Total" and variable1.get()=="Class":
                items=entry3.get().split(',')
                for item in items:
                    cls.append(item)
                if variable1.get()+"="+entry3.get() not in tags:
                    tags.append(variable1.get()+"="+entry3.get())
                if variable1.get()!="Select Filter" and entry3.get()!="":
                    set_tags(var1.get())
                entry3.delete(0,'end')
                variable1.set("Select Filter")
                log(0)
                return

            elif variable1.get()!="Select Filter":
                if entry3.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    sdf = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
                    sdf_rows = sdf.to_numpy().tolist()
                            
                    txt=entry3.get().lower().split(",")
                    tmp_rows=[]
                    tmp_rows1=df_rows
                    for item in txt:
                        for value in tmp_rows1:
                            for data in sdf_rows:
                                if data[26]==value[0]:
                                    cls_val=data[OptionList1[variable1.get()]]
                            if str(cls_val).lower()!=item.lower():
                                if var1.get()== 1:
                                    continue
                                else:
                                    tmp_rows1=list(filter(partial(ne, value), tmp_rows1))
                            else:
                                if var1.get()== 1:
                                    tmp_rows1=list(filter(partial(ne, value), tmp_rows1))
                                else:
                                    continue
                        if var1.get()==1:
                            tmp_rows=tmp_rows1
                        elif var1.get()==0:
                            tmp_rows=tmp_rows+tmp_rows1 
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    count=0
                    for row in tmp_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable1.get()+"="+entry3.get() not in tags:
                        tags.append(variable1.get()+"="+entry3.get())
                    set_tags(var1.get())
                    entry3.delete(0,'end')
                    variable1.set("Select Filter")
                    df_rows=tmp_rows
                    
                    return
        def remove_filter1():
            global df_rows
            global batch
            global tags
            global index
            entry3.delete(0,'end')
            entry1.delete(0,'end')
            entry2.delete(0,'end')
            log(0)
            v2.set("Tags: Entries")
            v1.set("Count: "+str(count))
            if tags:
                tags.pop()
            if not tags:
                batch=getLatestSheet()
                log(1)
            else:
                for items in tags:
                    data=items.split("=")
                    variable1.set(data[0])
                    entry3.insert(0,data[1])
                    excel_read()
            return

        def OnDoubleClick(event):
            
            def onclose():
                window2.grab_release()
                window2.destroy()
                return
            def update_det():
                global df_rows
                global batch
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    loc = ('data/studentlist.xlsx')
                    wb = xlrd.open_workbook(loc)
                    sheet = wb.sheet_by_name(batch)
                    sheet.cell_value(0, 0)
                    for i in range(1,sheet.nrows):
                        if sheet.cell_value(i, 26)==val[26]:
                            wb = load_workbook('./data/studentlist.xlsx')
                            ws = wb[batch]
                            ws[str(OptionList[variable.get()])+str(i+1)]=str(entry1.get())
                            wb.save('./data/studentlist.xlsx')
                            wb.close()
                            break
                        else:
                            if i==sheet.nrows:
                                messagebox.showwarning("Oops!","Student data not found..")
                                return
                            else:
                                continue   
                    messagebox.showinfo("Entry successful","Student Data has been saved successfully.")
                    batch=getLatestSheet()
                    df_rows=getdata()
                    state1()
                window2.grab_release()
                window2.destroy()
                return
            window2=Toplevel()
            window2.title("Update Student Details")
            window2.iconbitmap('data/icon.ico')
            w = 285 # width for the Tk root
            h = 100 # height for the Tk root
            ws = window2.winfo_screenwidth() # width of the screen
            hs = window2.winfo_screenheight() # height of the screen
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            window2.geometry('%dx%d+%d+%d' % (w, h, x, y))
            window2.resizable(width=False, height=False)
            my_canvas = Canvas(window2, width=285, height=100)
            my_canvas.pack(fill="both", expand=True)
            OptionList={"Class":"S","Roll No.":"T","First Name":"A","Last name":"B","Gender":"C","Age":"D","Father's Name":"E","Mother's Name":"F","Gaurdian's Name":"G","DOB":"H","Religion":"I","Address":"J","Father's Contact No.":"K","Mother's Contact No.":"L","Identity Mark":"M","Blood Group":"N","Community":"O","Tribe/Caste":"P"}
            variable = StringVar(window2)
            menu=OptionMenu(window2,variable,*OptionList)
            menu.config(width=15,anchor='w')
            menu_window = my_canvas.create_window(5, 10, anchor="nw", window=menu)
            variable.set("Roll No.")
            field1=Label(window2,text="*Note: select student attribute \nthat needs to be updated.",fg="grey")
            field1_window = my_canvas.create_window(5, 48, anchor="nw", window=field1)
            entry1=Entry(window2,width=20)
            entry1_window = my_canvas.create_window(150, 15, anchor="nw", window=entry1)
            button4 = Button(window2, text="SAVE", command=update_det)
            button4.config(height = 1, width=10)
            button4_window = my_canvas.create_window(190, 50, anchor="nw", window=button4)
            item = treeview.selection()
            val=treeview.item(item)["values"]
            window2.protocol("WM_DELETE_WINDOW", onclose)
            window2.grab_set()

            
        def g_reciept():
            global rp
            global batch
            font = ImageFont.truetype("arial.ttf", 20)
            shutil.copy2('data/reciept.png', 'data/tmp_reciept.png')
            img = Image.open('data/tmp_reciept.png')

            if state==5:
                wb = load_workbook('./data/studentlist.xlsx')
                if today.strftime("%Y") not in wb.sheetnames:
                    ws=wb.create_sheet(today.strftime("%Y"))
                    ws.append(['First Name','Last name','Gender','Age','Father\'s Name','Mother\'s Name','Gaurdian\'s Name','DOB','Religion','Address','Father\'s Contact No.','Mother\'s Contact No.','Identity Mark','Blood Group','Community','Tribe/Caste','Father\'s Occupation','Mother\'s Occupation','Class','Roll No.','Previous School','Result','Percentage','Prev. Attendance','Admission Date','Last Fee Payment Date','AdmissionID'
        ])
                wb.save('./data/studentlist.xlsx')
                wb.close()
                wb = xlrd.open_workbook('data/studentlist.xlsx')
                newsheet = wb.sheet_by_name(str(int(getLatestSheet())))
                newsheet.cell_value(0, 0)
                for i in range(1,newsheet.nrows):
                    if str(newsheet.cell_value(i, 26)).lower()==str(entry4.get()).lower():
                        messagebox.showwarning("Oops!","Admission already renewed.")
                        return
                wb = load_workbook('./data/studentlist.xlsx')
                if str(int(getLatestSheet())-1) in wb.sheetnames:
                    batch=int(getLatestSheet())-1
                    df_rows=getdata()
                else:
                    messagebox.showwarning("Oops!","No data found.")
                    wb.close()
                    return               
                wb.close()
                entry=[entry4.get(),entry21.get(),today.strftime("%d/%m/%Y"),entry22.get(),entry23.get()]
                for value in entry:
                    if value=="":
                        messagebox.showwarning("Oops!","Some fields are still empty.")
                        return
                flag=0
                for data in df_rows:
                    if str(data[26]).lower()==str(entry4.get()).lower():
                        d1 = ImageDraw.Draw(img)
                        d1.text((300, 225), (data[0]+" "+data[1]).upper(),font=font,fill=(0, 0, 0))
                        d1.text((270, 272),entry21.get(),font=font,fill=(0, 0, 0))
                        d1.text((600, 319), today.strftime("%d/%m/%Y"),font=font,fill=(0, 0, 0))
                        d1.text((320, 366), data[26],font=font,fill=(0, 0, 0))
                        flag=1
                        break
                if(flag==0):
                    messagebox.showwarning("Oops!","No data found.")
                    return
                    
            elif state==2:
                entry=[entry1.get(),entry2.get(),variable2.get(),entry4.get(),entry5.get(),entry6.get(),entry7.get(),entry8.get(),entry9.get(),entry10.get(),entry11.get(),entry12.get(),entry18.get(),entry13.get(),entry25.get(),variable.get(),entry15.get(),entry16.get(),entry17.get(),"0",entry19.get(),variable1.get(),entry22.get(),entry23.get(),today.strftime("%d/%m/%Y")," "]
                for value in entry:
                    if value=="":
                        messagebox.showwarning("Oops!","Some fields are still empty.")
                        return
                d1 = ImageDraw.Draw(img)
                d1.text((300, 225), (entry1.get()+" "+entry2.get()).upper(),font=font,fill=(0, 0, 0))
                d1.text((270, 272),entry17.get(),font=font,fill=(0, 0, 0))
                d1.text((600, 319), today.strftime("%d/%m/%Y"),font=font,fill=(0, 0, 0))
                d1.text((320, 366), adid(),font=font,fill=(0, 0, 0))
            img.save('data/tmp_reciept.png')
            resize()
            rp = tk.PhotoImage(file="data/sreciept.png")
            def onclose():
                window4.grab_release()
                window4.destroy()
                return
            window4 = Toplevel()
            window4.title("Payment Details")
            window4.iconbitmap('data/icon.ico')
            w = 800 # width for the Tk root
            h = 400 # height for the Tk root
            ws = window4.winfo_screenwidth() # width of the screen
            hs = window4.winfo_screenheight() # height of the screen
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            window4.geometry('%dx%d+%d+%d' % (w, h, x, y))
            window4.resizable(width=False, height=False)
            my_canvas = Canvas(window4, width=285, height=100)
            my_canvas.pack(fill="both", expand=True)
            my_canvas.create_image(-500,-500, image=bg, anchor="nw")
            r_canvas = Canvas(window4, width=400, height=380)
            r_canvas.create_image(28,2, image=rp, anchor="nw")
            r_canvas_window = my_canvas.create_window(390, 10, anchor="nw", window=r_canvas)
            
            
            field24=Label(window4,text="Admission Amount",font=txtfont2,bg='#ffffff')
            field24_window = my_canvas.create_window(30, 10, anchor="nw", window=field24)
            entry24=Lotfi(window4,width=20,font=txtfont3)
            entry24_window = my_canvas.create_window(30, 40, anchor="nw", window=entry24)
            

            field28=Label(window4,text="Uniform",font=txtfont2,bg='#ffffff')
            field28_window = my_canvas.create_window(30, 90, anchor="nw", window=field28)
            entry28=Lotfi(window4,width=20,font=txtfont3)
            entry28_window = my_canvas.create_window(30, 120, anchor="nw", window=entry28)
            
            field26=Label(window4,text="Other",font=txtfont2,bg='#ffffff')
            field26_window = my_canvas.create_window(30, 250, anchor="nw", window=field26)
            entry26=Lotfi(window4,width=20,font=txtfont3)
            entry26_window = my_canvas.create_window(30, 280, anchor="nw", window=entry26)

            field29=Label(window4,text="Books And Stationaries",font=txtfont2,bg='#ffffff')
            field29_window = my_canvas.create_window(30, 170, anchor="nw", window=field29)
            entry29=Lotfi(window4,width=20,font=txtfont3)
            entry29_window = my_canvas.create_window(30, 200, anchor="nw", window=entry29)

            field27=Label(window4,text="Description*",font=txtfont2,bg='#ffffff')
            field27_window = my_canvas.create_window(230, 10, anchor="nw", window=field27)
            entry27=Text(window4,width=16,height=8,font=txtfont3,bg="#fefefe")
            entry27_window = my_canvas.create_window(230, 40, anchor="nw", window=entry27)


            button5 = Button(window4, text="NEXT", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command= lambda: save_reciept(window4,entry24.get(),entry28.get(),entry26.get(),entry29.get(),entry27.get(1.0, "end-1c")))
            button5.config(height = 2, width=15)
            button5_window = my_canvas.create_window(30,345, anchor="nw", window=button5)
            
            field1=Label(window4,text="*Note: enter the amount for\nthe respective items\n and provide item details\nin the description box.",bg='#ffffff',fg="grey")
            field1_window = my_canvas.create_window(230, 200, anchor="nw", window=field1)

            window4.protocol("WM_DELETE_WINDOW", onclose)
            window4.grab_set()


        def save_reciept(window,admission,uniform,other,books,des):
            global rp
            flag=0
            fee=0
            if state==2:
                if(to_excel(admission,uniform,other,books,des)):
                    flag=1
            if state==3:
                if(to_excel1()):
                    flag=1
                    fee=1
            elif state==5:
                if(to_excel2(admission,uniform,other,books,des)):
                    flag=1
            if flag==1:
                window.grab_release()
                window.destroy()
                font = ImageFont.truetype("arial.ttf", 25)
                img = Image.open('data/tmp_reciept.png')
                if fee==0:
                    d1 = ImageDraw.Draw(img)
                    x1=145
                    x2=475
                    y=500
                    d1.text((x1, y),"ADMISSION",font=font,fill=(0, 0, 0))
                    d1.text((x2, y),admission,font=font,fill=(0, 0, 0))
                    d1.text((x1, y+40),"UNIFORM",font=font,fill=(0, 0, 0))
                    d1.text((x2, y+40),uniform,font=font,fill=(0, 0, 0))
                    d1.text((x1, y+80),"BOOKS/STATIONARIES",font=font,fill=(0, 0, 0))
                    d1.text((x2, y+80),books,font=font,fill=(0, 0, 0))
                    d1.text((x1, y+120),"OTHER",font=font,fill=(0, 0, 0))
                    d1.text((x2, y+120),other,font=font,fill=(0, 0, 0))
                    d1.text((x1, y+200),"DESCRIPTION:",font=font,fill=(0, 0, 0))
                    d1.text((x1+30, y+240),des,font=font,fill=(0, 0, 0))
                    img.save('data/tmp_reciept.png')
                    resize()
                rp = tk.PhotoImage(file="data/sreciept.png")
                def onclose():
                    window5.grab_release()
                    window5.destroy()
                    return
                window5 = Toplevel()
                window5.title("Payment Details")
                window5.iconbitmap('data/icon.ico')
                w = 400 # width for the Tk root
                h = 445 # height for the Tk root
                ws = window5.winfo_screenwidth() # width of the screen
                hs = window5.winfo_screenheight() # height of the screen
                x = (ws/2) - (w/2)
                y = (hs/2) - (h/2)
                window5.geometry('%dx%d+%d+%d' % (w, h, x, y))
                window5.resizable(width=False, height=False)
                my_canvas = Canvas(window5, width=285, height=100)
                my_canvas.pack(fill="both", expand=True)
                my_canvas.create_image(-500,-500, image=bg, anchor="nw")
                r_canvas = Canvas(window5, width=400, height=380)
                r_canvas.create_image(28,2, image=rp, anchor="nw")
                r_canvas_window = my_canvas.create_window(0, 10, anchor="nw", window=r_canvas)
                

                button6 = Button(window5, text="SAVE", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command = lambda:save_file(img))
                button6.config(height = 2, width=15)
                button6_window = my_canvas.create_window(135,400, anchor="nw", window=button6)

                window5.protocol("WM_DELETE_WINDOW", onclose)
                window5.grab_set()
                
        def saved_reciept(a):
            if variable.get()!="Total" and variable.get()!="Fee Status":
                global rp
                curitem=treeview.focus()
                row=treeview.item(curitem)['values']
                if variable.get()=="Fee Entries":
                    year=row[6][6:]
                    sdf = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=year)
                    sdf_rows = sdf.to_numpy().tolist()
                    for data in sdf_rows:
                        if data[26]==row[0]:
                            fname=data[0]
                            lname=data[1]
                            clss=data[18]
                            date=row[6]
                            amnt=row[1]
                            bamnt=row[2]
                            uamnt=row[3]
                            oamnt=row[4]
                            des=row[5]
                            label="FEE AMOUNT"
                elif variable.get()=="Admissions":
                    year=row[2][6:]
                    sdf = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=year)
                    sdf_rows = sdf.to_numpy().tolist()
                    for data in sdf_rows:
                        if data[26]==row[0]:
                            fname=data[0]
                            lname=data[1]
                            clss=data[18]
                            date=row[2]
                            amnt=row[1]
                            bamnt=row[3]
                            uamnt=row[4]
                            oamnt=row[5]
                            des=row[6]
                            label="ADMISSION"
                    
                font = ImageFont.truetype("arial.ttf", 25)
                shutil.copy2('data/reciept.png', 'data/tmp_reciept.png')
                img = Image.open('data/tmp_reciept.png')
                d1 = ImageDraw.Draw(img)
                x1=145
                x2=475
                y=500
                d1.text((300, 225), (fname+" "+lname).upper(),font=font,fill=(0, 0, 0))
                d1.text((270, 272), str(clss),font=font,fill=(0, 0, 0))
                d1.text((600, 319), date,font=font,fill=(0, 0, 0))
                d1.text((320, 366), row[0],font=font,fill=(0, 0, 0))
                d1.text((x1, y),label,font=font,fill=(0, 0, 0))
                d1.text((x2, y),str(amnt),font=font,fill=(0, 0, 0))
                d1.text((x1, y+40),"UNIFORM",font=font,fill=(0, 0, 0))
                d1.text((x2, y+40),str(uamnt),font=font,fill=(0, 0, 0))
                d1.text((x1, y+80),"BOOKS/STATIONARIES",font=font,fill=(0, 0, 0))
                d1.text((x2, y+80),str(bamnt),font=font,fill=(0, 0, 0))
                d1.text((x1, y+120),"OTHER",font=font,fill=(0, 0, 0))
                d1.text((x2, y+120),str(oamnt),font=font,fill=(0, 0, 0))
                d1.text((x1, y+200),"DESCRIPTION:",font=font,fill=(0, 0, 0))
                d1.text((x1+30, y+240),des,font=font,fill=(0, 0, 0))
                img.save('data/tmp_reciept.png')
                resize()
                rp = tk.PhotoImage(file="data/sreciept.png")
                def onclose():
                    window5.grab_release()
                    window5.destroy()
                    return
                window5 = Toplevel()
                window5.title("Payment Details")
                window5.iconbitmap('data/icon.ico')
                w = 400 # width for the Tk root
                h = 445 # height for the Tk root
                ws = window5.winfo_screenwidth() # width of the screen
                hs = window5.winfo_screenheight() # height of the screen
                x = (ws/2) - (w/2)
                y = (hs/2) - (h/2)
                window5.geometry('%dx%d+%d+%d' % (w, h, x, y))
                window5.resizable(width=False, height=False)
                my_canvas = Canvas(window5, width=285, height=100)
                my_canvas.pack(fill="both", expand=True)
                my_canvas.create_image(-500,-500, image=bg, anchor="nw")
                r_canvas = Canvas(window5, width=400, height=380)
                r_canvas.create_image(28,2, image=rp, anchor="nw")
                r_canvas_window = my_canvas.create_window(0, 10, anchor="nw", window=r_canvas)
                

                button6 = Button(window5, text="SAVE", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command = lambda:save_file(img))
                button6.config(height = 2, width=15)
                button6_window = my_canvas.create_window(135,400, anchor="nw", window=button6)

                window5.protocol("WM_DELETE_WINDOW", onclose)
                window5.grab_set()
                
        def f_reciept():
            global rp
            font = ImageFont.truetype("arial.ttf", 25)
            shutil.copy2('data/reciept.png', 'data/tmp_reciept.png')
            img = Image.open('data/tmp_reciept.png')
            d1 = ImageDraw.Draw(img)
            x1=145
            x2=475
            y=500


            entry=[entry1.get(),entry4.get(),entry5.get(),entry8.get(),entry6.get(),entry7.get(1.0, "end-1c")]
            for value in entry:
                if value=="" or variable.get()=="Month":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                
            wb = load_workbook('./data/studentlist.xlsx')
            batch=getLatestSheet()
            df_rows=getdata()           
            wb.close()
            
            flag=0
            for data in df_rows:
                if str(data[26]).lower()==str(entry1.get()).lower():
                    d1.text((300, 225), (data[0]+" "+data[1]).upper(),font=font,fill=(0, 0, 0))
                    d1.text((270, 272),str(data[18]),font=font,fill=(0, 0, 0))
                    d1.text((600, 319),today.strftime("%d/%m/%Y"),font=font,fill=(0, 0, 0))
                    d1.text((320, 366),data[26],font=font,fill=(0, 0, 0))
                    flag=1
                    break
            if(flag==0):
                messagebox.showwarning("Oops!","No data found2.")
                return
            
            d1.text((x1, y),"FEE AMOUNT",font=font,fill=(0, 0, 0))
            d1.text((x2, y),entry4.get(),font=font,fill=(0, 0, 0))
            d1.text((x1, y+40),"UNIFORM",font=font,fill=(0, 0, 0))
            d1.text((x2, y+40),entry8.get(),font=font,fill=(0, 0, 0))
            d1.text((x1, y+80),"BOOKS/STATIONARIES",font=font,fill=(0, 0, 0))
            d1.text((x2, y+80),entry5.get(),font=font,fill=(0, 0, 0))
            d1.text((x1, y+120),"OTHER",font=font,fill=(0, 0, 0))
            d1.text((x2, y+120),entry6.get(),font=font,fill=(0, 0, 0))
            d1.text((x1, y+200),"DESCRIPTION:",font=font,fill=(0, 0, 0))
            d1.text((x1+30, y+240),entry7.get(1.0, "end-1c"),font=font,fill=(0, 0, 0))
            img.save('data/tmp_reciept.png')
            resize()
            rp = tk.PhotoImage(file="data/sreciept.png")
            def onclose():
                window5.grab_release()
                window5.destroy()
                return
            window5 = Toplevel()
            window5.title("Payment Details")
            window5.iconbitmap('data/icon.ico')
            w = 400 # width for the Tk root
            h = 445 # height for the Tk root
            ws = window5.winfo_screenwidth() # width of the screen
            hs = window5.winfo_screenheight() # height of the screen
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            window5.geometry('%dx%d+%d+%d' % (w, h, x, y))
            window5.resizable(width=False, height=False)
            my_canvas = Canvas(window5, width=285, height=100)
            my_canvas.pack(fill="both", expand=True)
            my_canvas.create_image(-500,-500, image=bg, anchor="nw")
            r_canvas = Canvas(window5, width=400, height=380)
            r_canvas.create_image(28,2, image=rp, anchor="nw")
            r_canvas_window = my_canvas.create_window(0, 10, anchor="nw", window=r_canvas)
            

            button6 = Button(window5, text="NEXT", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command= lambda: save_reciept(window5,"","","","",""))
            button6.config(height = 2, width=15)
            button6_window = my_canvas.create_window(135,400, anchor="nw", window=button6)

            window5.protocol("WM_DELETE_WINDOW", onclose)
            window5.grab_set()

        def log_call():
            v2.set("Tags: Entries")
            log(1)

            
        if state == 2:
            OptionList=['ST','SC','OBC','General']
            variable = StringVar(window3)
            variable.set("General")
            OptionList1=['Pass','Fail']
            variable1 = StringVar(window3)
            variable1.set("Pass")
            OptionList2=['Male','Female','Other']
            variable2 = StringVar(window3)
            variable2.set("Male")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)


            section1=Label(text="New Admission",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 80, anchor="nw", window=section1)

            field1=Label(text="First Name",font=txtfont2,bg='#ffffff')
            field1_window = my_canvas.create_window(366, 160, anchor="nw", window=field1)
            entry1=Entry(window3,width=20,font=txtfont3)
            entry1_window = my_canvas.create_window(370, 200, anchor="nw", window=entry1)

            field2=Label(text="Last Name",font=txtfont2,bg='#ffffff')
            field2_window = my_canvas.create_window(366, 260, anchor="nw", window=field2)
            entry2=Entry(window3,width=20,font=txtfont3)
            entry2_window = my_canvas.create_window(370, 300, anchor="nw", window=entry2)

            field3=Label(text="Gender",font=txtfont2,bg='#ffffff')
            field3_window = my_canvas.create_window(366, 360, anchor="nw", window=field3)
            menu2=OptionMenu(window3,variable2,*OptionList2)
            menu2.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window2 = my_canvas.create_window(370, 400, anchor="nw", window=menu2)

            field4=Label(text="Age",font=txtfont2,bg='#ffffff')
            field4_window = my_canvas.create_window(366, 460, anchor="nw", window=field4)
            entry4=Lotfi(window3,width=20,font=txtfont3)
            entry4_window = my_canvas.create_window(370, 500, anchor="nw", window=entry4)

            field5=Label(text="Father's Name",font=txtfont2,bg='#ffffff')
            field5_window = my_canvas.create_window(366, 560, anchor="nw", window=field5)
            entry5=Entry(window3,width=20,font=txtfont3)
            entry5_window = my_canvas.create_window(370, 600, anchor="nw", window=entry5)

            field6=Label(text="Mother's Name",font=txtfont2,bg='#ffffff')
            field6_window = my_canvas.create_window(366, 660, anchor="nw", window=field6)
            entry6=Entry(window3,width=20,font=txtfont3)
            entry6_window = my_canvas.create_window(370, 700, anchor="nw", window=entry6)




            field7=Label(text="Gaurdian's Name",font=txtfont2,bg='#ffffff')
            field7_window = my_canvas.create_window(666, 160, anchor="nw", window=field7)
            entry7=Entry(width=20,font=txtfont3)
            entry7_window = my_canvas.create_window(670, 200, anchor="nw", window=entry7)

            field8=Label(text="Date Of Birth (dd/mm/yy)",font=txtfont2,bg='#ffffff')
            field8_window = my_canvas.create_window(666, 260, anchor="nw", window=field8)
            entry8=Entry(window3,width=20,font=txtfont3)
            entry8_window = my_canvas.create_window(670, 300, anchor="nw", window=entry8)

            field9=Label(text="Religion",font=txtfont2,bg='#ffffff')
            field9_window = my_canvas.create_window(666, 360, anchor="nw", window=field9)
            entry9=Entry(window3,width=20,font=txtfont3)
            entry9_window = my_canvas.create_window(670, 400, anchor="nw", window=entry9)

            field10=Label(text="Address",font=txtfont2,bg='#ffffff')
            field10_window = my_canvas.create_window(666, 460, anchor="nw", window=field10)
            entry10=Entry(window3,width=20,font=txtfont3)
            entry10_window = my_canvas.create_window(670, 500, anchor="nw", window=entry10)

            field11=Label(text="Father's Contact No.",font=txtfont2,bg='#ffffff')
            field11_window = my_canvas.create_window(666, 560, anchor="nw", window=field11)
            entry11=Lotfi(window3,width=20,font=txtfont3)
            entry11_window = my_canvas.create_window(670, 600, anchor="nw", window=entry11)

            field12=Label(text="Mother's Contact No.",font=txtfont2,bg='#ffffff')
            field12_window = my_canvas.create_window(666, 660, anchor="nw", window=field12)
            entry12=Lotfi(window3,width=20,font=txtfont3)
            entry12_window = my_canvas.create_window(670, 700, anchor="nw", window=entry12)




            field18=Label(text="Idendity Mark",font=txtfont2,bg='#ffffff')
            field18_window = my_canvas.create_window(966, 160, anchor="nw", window=field18)
            entry18=Entry(window3,width=20,font=txtfont3)
            entry18_window = my_canvas.create_window(970, 200, anchor="nw", window=entry18)

            field13=Label(text="Blood Group",font=txtfont2,bg='#ffffff')
            field13_window = my_canvas.create_window(966, 260, anchor="nw", window=field13)
            entry13=Entry(window3,width=20,font=txtfont3)
            entry13_window = my_canvas.create_window(970, 300, anchor="nw", window=entry13)

            field14=Label(text="Tribe/Caste",font=txtfont2,bg='#ffffff')
            field14_window = my_canvas.create_window(966, 360, anchor="nw", window=field14)

            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(970, 400, anchor="nw", window=menu)


            field25=Label(text="Community",font=txtfont2,bg='#ffffff')
            field25_window = my_canvas.create_window(966, 460, anchor="nw", window=field25)
            entry25=Entry(window3,width=20,font=txtfont3)
            entry25_window = my_canvas.create_window(970, 500, anchor="nw", window=entry25)
            
            field15=Label(text="Father's Occupation",font=txtfont2,bg='#ffffff')
            field15_window = my_canvas.create_window(966, 560, anchor="nw", window=field15)
            entry15=Entry(window3,width=20,font=txtfont3)
            entry15_window = my_canvas.create_window(970, 600, anchor="nw", window=entry15)

            field16=Label(text="Mother's Occupation",font=txtfont2,bg='#ffffff')
            field16_window = my_canvas.create_window(966, 660, anchor="nw", window=field16)
            entry16=Entry(window3,width=20,font=txtfont3)
            entry16_window = my_canvas.create_window(970, 700, anchor="nw", window=entry16)

            field17=Label(text="Class",font=txtfont2,bg='#ffffff')
            field17_window = my_canvas.create_window(1266, 160, anchor="nw", window=field17)
            entry17=Entry(window3,width=20,font=txtfont3)
            entry17_window = my_canvas.create_window(1270, 200, anchor="nw", window=entry17)

            field19=Label(text="Previous School",font=txtfont2,bg='#ffffff')
            field19_window = my_canvas.create_window(1266, 260, anchor="nw", window=field19)
            entry19=Entry(window3,width=20,font=txtfont3)
            entry19_window = my_canvas.create_window(1270, 300, anchor="nw", window=entry19)

            field21=Label(text="Result",font=txtfont2,bg='#ffffff')
            field21_window = my_canvas.create_window(1266, 360, anchor="nw", window=field21)
            menu1=OptionMenu(window3,variable1,*OptionList1)
            menu1.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window1 = my_canvas.create_window(1270, 400, anchor="nw", window=menu1)

            field22=Label(text="Result Percentage(%)",font=txtfont2,bg='#ffffff')
            field22_window = my_canvas.create_window(1266, 460, anchor="nw", window=field22)
            entry22=Entry(window3,width=20,font=txtfont3)
            entry22_window = my_canvas.create_window(1270, 500, anchor="nw", window=entry22)

            
            field23=Label(text="Attendance",font=txtfont2,bg='#ffffff')
            field23_window = my_canvas.create_window(1266, 560, anchor="nw", window=field23)
            entry23=Entry(window3,width=20,font=txtfont3)
            entry23_window = my_canvas.create_window(1270, 600, anchor="nw", window=entry23)

            button4 = Button(window3, text="NEXT", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command= g_reciept)
            button4.config(height = 2, width=20)
            button4_window = my_canvas.create_window(1270, 680, anchor="nw", window=button4)

        elif state == 3:
            OptionList={'January':'B','February':'C','March':'D','April':'E','May':'F','June':'G','July':'H','August':'I','September':'J','October':'K','November':'L','December':'M'}
            variable = StringVar(window3)
            variable.set("Month")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)

            section1=Label(text="Fee Entry",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)

            field1=Label(text="Admission ID",font=txtfont2,bg='#ffffff')
            field1_window = my_canvas.create_window(366, 180, anchor="nw", window=field1)
            entry1=Entry(width=20,font=txtfont3)
            entry1_window = my_canvas.create_window(370, 220, anchor="nw", window=entry1)

            field9=Label(text="Paying For",font=txtfont2,bg='#ffffff')
            field9_window = my_canvas.create_window(366, 280, anchor="nw", window=field9)

            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(370, 320, anchor="nw", window=menu)
            
            button3 = Button(window3, text="Add", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=addmonth)
            button3.config(height = 2, width=8)
            button3_window = my_canvas.create_window(540, 316, anchor="nw", window=button3)
            
            field4=Label(text="Fee Amount",font=txtfont2,bg='#ffffff')
            field4_window = my_canvas.create_window(366, 380, anchor="nw", window=field4)
            entry4=Lotfi(width=20,font=txtfont3)
            entry4_window = my_canvas.create_window(370, 420, anchor="nw", window=entry4)
            
            field5=Label(text="Books And Stationaries",font=txtfont2,bg='#ffffff')
            field5_window = my_canvas.create_window(666, 180, anchor="nw", window=field5)
            entry5=Lotfi(width=20,font=txtfont3)
            entry5_window = my_canvas.create_window(670, 220, anchor="nw", window=entry5)        

            field8=Label(text="Uniform",font=txtfont2,bg='#ffffff')
            field8_window = my_canvas.create_window(666, 280, anchor="nw", window=field8)
            entry8=Lotfi(width=20,font=txtfont3)
            entry8_window = my_canvas.create_window(670, 320, anchor="nw", window=entry8)
            
            field6=Label(text="Other",font=txtfont2,bg='#ffffff')
            field6_window = my_canvas.create_window(666, 380, anchor="nw", window=field6)
            entry6=Lotfi(window3,width=20,font=txtfont3)
            entry6_window = my_canvas.create_window(670, 420, anchor="nw", window=entry6)
            
            field7=Label(text="Description*",font=txtfont2,bg='#ffffff')
            field7_window = my_canvas.create_window(966, 180, anchor="nw", window=field7)
            entry7=Text(window3,width=20,height=6,font=txtfont3,bg="#fefefe")
            entry7_window = my_canvas.create_window(970, 220, anchor="nw", window=entry7)

            section2=Label(text="For the month of",textvariable=u,font=txtfont,bg='#ffffff',fg="grey")
            section2_window = my_canvas.create_window(368, 470, anchor="nw", window=section2)
            
            button4 = Button(window3, text="NEXT", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=f_reciept)
            button4.config(height = 2, width=20)
            button4_window = my_canvas.create_window(370, 500, anchor="nw", window=button4)
            
        elif state == 1:
            global count
            global tags
            tags=[]
            count=0
            OptionList={"Fee Status":25,"AdmissionID":26,"Class":18,"Roll No.":19,"Batch":0,"First Name":0,"Last name":1,"Gender":2,"Age":3,"Community":14,"Tribe/Caste":15,"Religion":8,"Father's Name":4,"Mother's Name":5,"Gaurdian's Name":6,"DOB":7,"Address":9,"Father's Contact No.":10,"Mother's Contact No.":11,"Identity Mark":12,"Blood Group":13,"Prev. School":20,"Result":21}
            variable = StringVar(window3)
            variable.set("Select Filter")
            button5 = Button(window3, text="Student List", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)
            
            section1=Label(text="Student List",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)

            section2=Label(text="Tags: Students",textvariable=v,font=txtfont,bg='#ffffff',fg="grey")
            section2_window = my_canvas.create_window(245, 275, anchor="nw", window=section2)

            section3=Label(text="Count:",textvariable=v1,font=txtfont,bg='#ffffff',fg="grey")
            section3_window = my_canvas.create_window(1400, 275, anchor="nw", window=section3)
            
            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(355, 180, anchor="nw", window=menu)

            c1 = tk.Checkbutton(window3, text='Reverse',variable=var1, onvalue=1, offvalue=0,bg='#ffffff')
            check_window = my_canvas.create_window(355, 220, anchor="nw", window=c1)

            entry1=Entry(window3,width=11,font=txtfont4)
            entry1_window = my_canvas.create_window(540, 182, anchor="nw", window=entry1)        

            button4 = Button(window3, text="Add Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=rd_excel)
            button4.config(height = 2, width=10)
            button4_window = my_canvas.create_window(680, 177, anchor="nw", window=button4)

            button5 = Button(window3, text="Remove Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=remove_filter)
            button5.config(height = 2, width=12)
            button5_window = my_canvas.create_window(770, 177, anchor="nw", window=button5)

            button6 = Button(window3, text="Assign Roll", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=assign_roll)
            button6.config(height = 2, width=13)
            button6_window = my_canvas.create_window(1370, 177, anchor="nw", window=button6)
            
            button7 = Button(window3, text="Update Data", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=update_data)            
            button7.config(height = 2, width=13)
            button7_window = my_canvas.create_window(1370, 220, anchor="nw", window=button7)

            
            tree_frame = Frame(window3,width=1289, height=500)
            tree_frame_window=my_canvas.create_window(245, 298, anchor="nw", window=tree_frame)
            tree_scrolly = Scrollbar(tree_frame)
            tree_scrolly.pack(side=RIGHT, fill=Y)
            tree_frame.pack_propagate(0)
            treeview=ttk.Treeview(tree_frame, yscrollcommand=tree_scrolly.set, selectmode="extended")
            treeview.pack(side="top", fill="both", expand=True)
            scroll = Scrollbar(tree_frame, orient=HORIZONTAL, command=treeview.xview)
            treeview.configure(xscrollcommand=scroll.set)
            tree_scrolly.config(command=treeview.yview)
            scroll.pack(side="bottom", fill="x")
            treeview["column"] = list(df.columns)
            treeview["show"] = "headings"
            treeview.tag_configure('oddrow', background="white")
            treeview.tag_configure('evenrow', background="#F7F7F7")
            for column in treeview["column"]:
                treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                 treeview_sort_column(treeview, _column, False))
            for row in df_rows:
                if count % 2 == 0:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                else:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                count=count+1
            v1.set("Count: "+str(count))
            treeview.bind("<Double-1>", OnDoubleClick)
        elif state == 4:
            OptionList=["Admissions","Fee Entries","Fee Status","Total"]
            variable = StringVar(window3)
            variable.set("Total")
            OptionList1={"Batch":0,"Class":18,"AdmissionID":26}
            variable1 = StringVar(window3)
            variable1.set("Select Filter")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)

            section1=Label(text="Entry Log",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)

            section2=Label(text="Tags: Entries",textvariable=v2,font=txtfont,bg='#ffffff',fg="grey")
            section2_window = my_canvas.create_window(245, 275, anchor="nw", window=section2)

            section3=Label(text="Count :",textvariable=v1,font=txtfont,bg='#ffffff',fg="grey")
            section3_window = my_canvas.create_window(1400, 275, anchor="nw", window=section3)

            field1=Label(text="Date (dd/mm/yy)",font=txtfont,bg='#ffffff',fg="grey")
            field1_window = my_canvas.create_window(355, 156, anchor="nw", window=field1)
            entry1=Entry(window3,width=11,font=txtfont4)
            entry1_window = my_canvas.create_window(355, 180, anchor="nw", window=entry1)
            field2=Label(text="To",font=txtfont2,bg='#ffffff')
            field2_window = my_canvas.create_window(497, 180, anchor="nw", window=field2)
            entry2=Entry(window3,width=11,font=txtfont4)
            entry2_window = my_canvas.create_window(540, 180, anchor="nw", window=entry2)

            button4 = Button(window3, text="Search", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=log_call)
            button4.config(height = 2, width=10)
            button4_window = my_canvas.create_window(680, 177, anchor="nw", window=button4)

            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=12,anchor='w')
            menu_window = my_canvas.create_window(354, 220, anchor="nw", window=menu)

            menu1=OptionMenu(window3,variable1,*OptionList1)
            menu1.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu1_window = my_canvas.create_window(905, 180, anchor="nw", window=menu1)

            c1 = tk.Checkbutton(window3, text='Reverse',variable=var1, onvalue=1, offvalue=0, bg='#ffffff')
            check_window = my_canvas.create_window(905, 220, anchor="nw", window=c1)

            entry3=Entry(window3,width=11,font=txtfont4)
            entry3_window = my_canvas.create_window(1090, 182, anchor="nw", window=entry3)        

            button6 = Button(window3, text="Add Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=excel_read)
            button6.config(height = 2, width=10)
            button6_window = my_canvas.create_window(1230, 177, anchor="nw", window=button6)

            button5 = Button(window3, text="Remove Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=remove_filter1)
            button5.config(height = 2, width=12)
            button5_window = my_canvas.create_window(1320, 177, anchor="nw", window=button5)
            
            tree_frame = Frame(window3,width=1289, height=500)
            tree_frame_window=my_canvas.create_window(245, 298, anchor="nw", window=tree_frame)
            tree_scrolly = Scrollbar(tree_frame)
            tree_scrolly.pack(side=RIGHT, fill=Y)
            tree_frame.pack_propagate(0)
            treeview=ttk.Treeview(tree_frame, yscrollcommand=tree_scrolly.set, selectmode="extended")
            treeview.pack(side="top", fill="both", expand=True)
            scroll = Scrollbar(tree_frame, orient=HORIZONTAL, command=treeview.xview)
            treeview.configure(xscrollcommand=scroll.set)
            tree_scrolly.config(command=treeview.yview)
            scroll.pack(side="bottom", fill="x")
            log(1)
            treeview.bind("<Double-1>", saved_reciept)
            
        elif state == 5:
            OptionList=['Pass','Fail']
            variable = StringVar(window3)
            variable.set("Pass")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)


            section1=Label(text="Renew Admission",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)
            
            field4=Label(text="Admission ID",font=txtfont2,bg='#ffffff')
            field4_window = my_canvas.create_window(366, 180, anchor="nw", window=field4)
            entry4=Entry(width=20,font=txtfont3)
            entry4_window = my_canvas.create_window(370, 220, anchor="nw", window=entry4)
            
            field6=Label(text="Result",font=txtfont2,bg='#ffffff')
            field6_window = my_canvas.create_window(366, 280, anchor="nw", window=field6)
            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(370, 320, anchor="nw", window=menu)

            field22=Label(text="Result Percentage(%)",font=txtfont2,bg='#ffffff')
            field22_window = my_canvas.create_window(666, 180, anchor="nw", window=field22)
            entry22=Entry(window3,width=20,font=txtfont3)
            entry22_window = my_canvas.create_window(670, 220, anchor="nw", window=entry22)

            
            field23=Label(text="Attendance",font=txtfont2,bg='#ffffff')
            field23_window = my_canvas.create_window(666, 280, anchor="nw", window=field23)
            entry23=Entry(window3,width=20,font=txtfont3)
            entry23_window = my_canvas.create_window(670, 320, anchor="nw", window=entry23)
            
            field21=Label(text="Class",font=txtfont2,bg='#ffffff')
            field21_window = my_canvas.create_window(666, 380, anchor="nw", window=field21)
            entry21=Entry(window3,width=20,font=txtfont3)
            entry21_window = my_canvas.create_window(670, 420, anchor="nw", window=entry21)

            button4 = Button(window3, text="NEXT", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=g_reciept)
            button4.config(height = 2, width=20)
            button4_window = my_canvas.create_window(370, 410, anchor="nw", window=button4)
    form()
def check_cred():
    if entry1.get()=="admin" and entry2.get()=="admin1234":
        start()
    else:
        messagebox.showwarning("Oops!","Invalid Credentials!")
        entry1.delete(0,'end')
        entry2.delete(0,'end')
        return

    return
    
window = Tk()
window.title("Admin Login")
window.iconbitmap('data/icon.ico')
w = 250 # width for the Tk root
h = 140 # height for the Tk root
ws = window.winfo_screenwidth() # width of the screen
hs = window.winfo_screenheight() # height of the screen
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)
window.geometry('%dx%d+%d+%d' % (w, h, x, y))
window.resizable(width=False, height=False)
my_canvas = Canvas(window, width=285, height=100)
my_canvas.pack(fill="both", expand=True)

tags=[]
months=[]
today = date.today()
state=1
state1=0
count=0
index=0
cls=[]
wb = load_workbook('./data/studentlist.xlsx')
batch = wb.sheetnames[len(wb.sheetnames)-1]
wb.close()
df = pd.read_excel('./data/studentlist.xlsx', dtype=str, keep_default_na=False,sheet_name=batch)
df_rows = df.to_numpy().tolist()
app_init = False

field1=Label(text="User Name*")
field1_window = my_canvas.create_window(20, 25, anchor="nw", window=field1)
entry1=Entry(width=20)
entry1_window = my_canvas.create_window(100, 25, anchor="nw", window=entry1)
field2=Label(text="Password*")
field2_window = my_canvas.create_window(20, 60, anchor="nw", window=field2)
entry2=Entry(width=20,show='*')
entry2_window = my_canvas.create_window(100, 60, anchor="nw", window=entry2)
button4 = Button(window, text="Login", command=check_cred)
button4.config(height = 1, width=10)
button4_window = my_canvas.create_window(102, 100, anchor="nw", window=button4)

class Lotfi(tk.Entry):
    def __init__(self, master=None, **kwargs):
        self.var = tk.StringVar()
        tk.Entry.__init__(self, master, textvariable=self.var, **kwargs)
        self.old_value = ''
        self.var.trace('w', self.check)
        self.get, self.set = self.var.get, self.var.set

    def check(self, *args):
        if self.get().isdigit(): 
            # the current value is only digits; allow this
            self.old_value = self.get()
        else:
            # there's non-digit characters in the input; reject this 
            self.set(self.old_value)

tk.mainloop()
